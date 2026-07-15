"""Excel reading service."""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Any


SUPPORTED_EXCEL_SUFFIXES = {".xlsx", ".xlsm", ".xltx", ".xltm"}


class ExcelServiceError(RuntimeError):
    """Raised when email keywords cannot be read from Excel."""


@dataclass(frozen=True)
class ExcelReadResult:
    """Structured result for Excel keyword extraction."""

    file_path: str
    sheet_name: str
    keywords: list[str]

    @property
    def total(self) -> int:
        return len(self.keywords)


class ExcelService:
    """Read email names from the first column of an Excel worksheet."""

    def read_mail_names(self, file_path: str, sheet_name: str | None = None) -> list[str]:
        """Return non-empty first-column values as clean strings."""

        return self.read(file_path, sheet_name).keywords

    def read(self, file_path: str, sheet_name: str | None = None) -> ExcelReadResult:
        """Read email names from the first column.

        The first worksheet is used when `sheet_name` is not provided.
        Empty rows are skipped. Values are converted to strings and stripped.
        """

        path = self._validate_file(file_path)
        load_workbook = self._import_load_workbook()

        workbook: Any | None = None
        try:
            workbook = load_workbook(
                filename=path,
                read_only=True,
                data_only=True,
            )
            worksheet = self._select_worksheet(workbook, sheet_name)
            keywords = self._read_first_column(worksheet)
        except ExcelServiceError:
            raise
        except Exception as exc:
            raise ExcelServiceError(f"读取Excel失败: {path}; {exc}") from exc
        finally:
            if workbook is not None:
                workbook.close()

        if not keywords:
            raise ExcelServiceError(f"Excel第一列没有可处理的邮件名称: {path}")

        return ExcelReadResult(
            file_path=str(path),
            sheet_name=str(worksheet.title),
            keywords=keywords,
        )

    def _validate_file(self, file_path: str) -> Path:
        path = Path(file_path).expanduser()

        if not path.exists():
            raise ExcelServiceError(f"Excel文件不存在: {path}")
        if not path.is_file():
            raise ExcelServiceError(f"Excel路径不是文件: {path}")
        if path.suffix.lower() not in SUPPORTED_EXCEL_SUFFIXES:
            supported = ", ".join(sorted(SUPPORTED_EXCEL_SUFFIXES))
            raise ExcelServiceError(f"不支持的Excel格式: {path.suffix}; 支持: {supported}")

        return path

    @staticmethod
    def _select_worksheet(workbook: Any, sheet_name: str | None) -> Any:
        if sheet_name:
            if sheet_name not in workbook.sheetnames:
                raise ExcelServiceError(f"Excel中不存在工作表: {sheet_name}")
            return workbook[sheet_name]

        if not workbook.sheetnames:
            raise ExcelServiceError("Excel文件中没有工作表")

        return workbook[workbook.sheetnames[0]]

    @staticmethod
    def _read_first_column(worksheet: Any) -> list[str]:
        keywords: list[str] = []

        for row in worksheet.iter_rows(min_row=1, max_col=1, values_only=True):
            if not row:
                continue

            value = row[0]
            if value is None:
                continue

            keyword = str(value).strip()
            if keyword:
                keywords.append(keyword)

        return keywords

    @staticmethod
    def _import_load_workbook() -> Any:
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise ExcelServiceError("未安装openpyxl，请先执行: pip install -r requirements.txt") from exc

        return load_workbook
